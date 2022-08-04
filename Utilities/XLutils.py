import openpyxl

def getrowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def getcolumncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)

def readdata(file,sheetname,rownum, columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(row=rownum,column=columnno).value

# def writedata(file,sheetname,rownum, columnno, data):
#     workbook=openpyxl.load_workbook(file)
#     sheet=workbook[sheetname]
#     return sheet.cell(row=rownum,column=columnno).value =data
#     workbook.save(file)
#
